from pandas import *
from utils.Statistics import get_statistics
from openpyxl import load_workbook
from utils.branch_wise_analysis import *
def value(X):
     if X=="A+":
          return 10
     elif X=="A":
          return 9
     elif X=="B":
          return 8
     elif X=="C":
          return 7
     elif X=="D":
          return 6
     elif X=="E":
          return 5
     elif X=="F" or X=="AB" or X=="ABSENT" or X=="MP":
          return 0

def reval_func(GPA_file, data):
    status=0
    flag=1
    try:
        civil=read_excel(GPA_file,sheet_name=["CE"])
        civil=DataFrame(civil["CE"])
    except:
        pass
    try:
        eee=read_excel(GPA_file,sheet_name=["EEE"])
        eee=DataFrame(eee["EEE"])
    except:
        pass
    try:
        mech=read_excel(GPA_file,sheet_name=["ME"])
        mech=DataFrame(mech["ME"])
    except:
        pass
    try:
        ece=read_excel(GPA_file,sheet_name=["ECE"])
        ece=DataFrame(ece["ECE"])
    except:
        pass
    try:
        cse=read_excel(GPA_file,sheet_name=["CSE"])
        cse=DataFrame(cse["CSE"])
    except:
        pass
    
    def change_data(df):
        for j in range(len(df)):
            if df.iloc[j,0] == data.iloc[i,0]:
                for k in df.columns:
                    if data.iloc[i,1] in k:
                        if df.loc[j,k] !=data.iloc[i,-2]:                        
                            if df.loc[j,k]=='A+' or df.loc[j,k]=='A' or df.loc[j,k]=='B' or df.loc[j,k]=='C' or df.loc[j,k]=='D' or df.loc[j,k]=='E':
                                df.iloc[j,-2]-=value(df.loc[j,k])*float(data.iloc[i,-1]) #Changing GPA
                                df.iloc[j,-7]-=value(df.loc[j,k])*10  #Changing GBM
                            df.loc[j,k]= data.iloc[i,-2]   #Modifying the grade
                            df.iloc[j,-2]+=value(data.iloc[i,-2])*float(data.iloc[i,-1])    #Changing GPA
                            df.iloc[j,-7]+=value(df.loc[j,k])*10  #Changing GBM                  
                            #print(df.iloc[j,-1])
                            df.iloc[j,-1]=df.iloc[j,-2]/df.iloc[j,-6]   #Modifying Points
                            #print(df.iloc[j,-1])
                            l=[]
                            for x in range(1,len(df.columns)-7):
                                l.append(df.iloc[j,x])   #Taking note of all grades
                            if "F" not in l and "AB" not in l and "ABSENT" not in l and "MP" not in l:
                                df.iloc[j,-5]="Pass"  #Modifying status column
                            df.iloc[j,-4]=l.count("F")+l.count("AB")+l.count("ABSENT")+l.count("MP")  #Modifying Backlogs count
                            df.iloc[j,-3]=df.iloc[j,-7]/len(l)-(l.count("COMPLE")+l.count("COMPLETED")) #Modifying pass percentage                        
        return df

    for i in range(len(data)):
        x=int(data.iloc[i,0][7:10])  
        try:  
            if x//100==1:
                if data.iloc[i,-2] !="No Change":
                    civil=change_data(civil)
        except:
            pass
        try:
            if x//100==2:
                if data.iloc[i,-2] !="No Change":
                    change_data(eee)
        except:
            pass
        try:
            if x//100==3:
                if data.iloc[i,-2] !="No Change":
                    change_data(mech)
        except:
            pass  
        try:
            if x//100==4:
                if data.iloc[i,-2] !="No Change":
                    change_data(ece)
        except:
            pass
        try:
            if x//100==5:
                if data.iloc[i,-2] !="No Change":
                    change_data(cse)
        except:
            pass
    
    with ExcelWriter(GPA_file,engine='openpyxl',mode='w') as output:
        try:
            civil.to_excel(output,sheet_name="CE",index=False)
        except:
            pass
        try:
            eee.to_excel(output,sheet_name="EEE",index=False)
        except:
            pass
        try:
            mech.to_excel(output,sheet_name="ME",index=False)
        except:
            pass
        try:
            ece.to_excel(output,sheet_name="ECE",index=False)
        except:
            pass
        try:
            cse.to_excel(output,sheet_name="CSE",index=False)
        except:
            pass
    get_statistics(GPA_file)
    wb=load_workbook(GPA_file)
    if "CE" not in wb.get_sheet_names() or "EEE" not in wb.get_sheet_names() or "ME" not in wb.get_sheet_names() or "ECE" not in wb.get_sheet_names() or "CSE" not in wb.get_sheet_names() or "CE stats" not in wb.get_sheet_names() or "EEE stats" not in wb.get_sheet_names() or "ME stats" not in wb.get_sheet_names() or "ECE stats" not in wb.get_sheet_names() or "CSE stats" not in wb.get_sheet_names():
        branch=wb.sheetnames[0]
        flag=0
    wb.save(GPA_file)
    if flag==0:
        branchwise_analysis(GPA_file,branch)